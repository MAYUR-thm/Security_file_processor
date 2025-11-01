"""
Excel Processor Module

This module handles processing of Excel files including:
- Text extraction from cells
- Structure preservation
- Cell formatting preservation
- PII detection and cleaning
"""

import logging
from pathlib import Path
from typing import Dict, List, Optional, Union, Any
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font
import json

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelProcessor:
    """Main class for processing Excel files."""

    def __init__(self):
        """Initialize the Excel processor."""
        self.sensitive_columns = {
            'employee full name': 'PERSON',
            'verified by (ra name)': 'PERSON',
            'ra signature / initials': 'PERSON',
            'authorized by (supervisor)': 'PERSON',
            'supervisor email': 'email'
        }

        # PII replacement patterns
        self.replacements = {
            'PERSON': '[REDACTED NAME]',
            'email': '[REDACTED EMAIL]',
            'phone': '[REDACTED PHONE]',
            'address': '[REDACTED ADDRESS]',
            'date': '[REDACTED DATE]'
        }

        # Style for redacted cells
        self.redacted_fill = PatternFill(start_color='FFFF0000', 
                                       end_color='FFFF0000', 
                                       fill_type='solid')
        self.redacted_font = Font(color='FFFFFF')  # White text

    def process_file(self, 
                    input_file: Union[str, Path], 
                    output_file: Optional[Union[str, Path]] = None,
                    mask_mode: str = "replace") -> Dict[str, Any]:
        """
        Process an Excel file, detecting and cleaning PII while preserving structure.
        
        Args:
            input_file: Path to input Excel file
            output_file: Path for output Excel file
            mask_mode: How to handle PII ('replace', 'mask', or 'remove')
            
        Returns:
            Dictionary containing processing results
        """
        input_file = Path(input_file)
        if not input_file.exists():
            raise FileNotFoundError(f"Input file not found: {input_file}")

        if not output_file:
            output_file = input_file.parent / f"cleaned_{input_file.name}"
        else:
            output_file = Path(output_file)

        try:
            # Read with pandas for data manipulation
            df = pd.read_excel(input_file)
            
            # Track changes for reporting
            stats = {
                'total_cells': df.size,
                'cells_with_pii': 0,
                'pii_types': {},
                'columns_cleaned': [],
                'success': False
            }

            # Process sensitive columns
            for column in df.columns:
                col_lower = column.lower()
                if col_lower in self.sensitive_columns:
                    pii_type = self.sensitive_columns[col_lower]
                    replacement = self.replacements.get(pii_type, '[REDACTED]')
                    
                    # Count non-null values as PII instances
                    pii_count = df[column].notna().sum()
                    stats['cells_with_pii'] += pii_count
                    stats['pii_types'][pii_type] = stats['pii_types'].get(pii_type, 0) + pii_count
                    stats['columns_cleaned'].append(column)
                    
                    # Apply redaction
                    if mask_mode == "remove":
                        df[column] = None
                    else:
                        df[column] = df[column].apply(
                            lambda x: replacement if pd.notna(x) else x
                        )

            # Save with pandas for data
            df.to_excel(output_file, index=False)

            # Re-open with openpyxl for formatting
            wb = openpyxl.load_workbook(output_file)
            ws = wb.active

            # Apply formatting to redacted cells
            for column in stats['columns_cleaned']:
                col_idx = None
                # Find column index
                for idx, cell in enumerate(ws[1], 1):
                    if cell.value == column:
                        col_idx = idx
                        break
                
                if col_idx:
                    for row in range(2, ws.max_row + 1):
                        cell = ws.cell(row=row, column=col_idx)
                        if cell.value and any(repl in str(cell.value) 
                                           for repl in self.replacements.values()):
                            cell.fill = self.redacted_fill
                            cell.font = self.redacted_font

            wb.save(output_file)
            
            stats['success'] = True
            stats['output_file'] = str(output_file)
            logger.info(f"Successfully processed Excel file: {input_file}")
            
            return stats

        except Exception as e:
            logger.error(f"Error processing Excel file {input_file}: {str(e)}")
            raise

    def extract_text(self, file_path: Union[str, Path]) -> str:
        """
        Extract text content from Excel file for analysis.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Extracted text content
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            df = pd.read_excel(file_path)
            text_parts = []

            # Add column headers
            text_parts.append("=== Column Headers ===")
            for col in df.columns:
                text_parts.append(f"Column: {col}")

            # Add cell contents
            text_parts.append("\n=== Cell Contents ===")
            for idx, row in df.iterrows():
                for col in df.columns:
                    if pd.notna(row[col]):
                        text_parts.append(f"{col}: {row[col]}")

            return "\n".join(text_parts)

        except Exception as e:
            logger.error(f"Error extracting text from Excel file {file_path}: {str(e)}")
            raise

    def analyze_structure(self, file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Analyze the structure of an Excel file.
        
        Args:
            file_path: Path to Excel file
            
        Returns:
            Dictionary containing structural analysis
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")

        try:
            df = pd.read_excel(file_path)
            
            analysis = {
                'total_rows': len(df),
                'total_columns': len(df.columns),
                'column_names': list(df.columns),
                'column_types': df.dtypes.to_dict(),
                'null_counts': df.isnull().sum().to_dict(),
                'potentially_sensitive_columns': []
            }

            # Identify potentially sensitive columns
            for column in df.columns:
                col_lower = column.lower()
                if any(term in col_lower for term in ['name', 'email', 'phone', 'address', 
                                                    'ssn', 'social', 'birth', 'employee']):
                    analysis['potentially_sensitive_columns'].append(column)

            return analysis

        except Exception as e:
            logger.error(f"Error analyzing Excel file {file_path}: {str(e)}")
            raise


def process_excel_file(input_file: Union[str, Path], 
                      output_file: Optional[Union[str, Path]] = None,
                      mask_mode: str = "replace") -> Dict[str, Any]:
    """
    Convenience function to process a single Excel file.
    
    Args:
        input_file: Path to input Excel file
        output_file: Path for output Excel file
        mask_mode: How to handle PII ('replace', 'mask', or 'remove')
        
    Returns:
        Processing results dictionary
    """
    processor = ExcelProcessor()
    return processor.process_file(input_file, output_file, mask_mode)


if __name__ == "__main__":
    # Example usage
    import sys
    
    if len(sys.argv) > 1:
        input_file = sys.argv[1]
        output_file = sys.argv[2] if len(sys.argv) > 2 else None
        
        try:
            processor = ExcelProcessor()
            
            # First analyze the structure
            print("\nAnalyzing Excel structure...")
            analysis = processor.analyze_structure(input_file)
            print("\nFile Structure:")
            print(f"Total rows: {analysis['total_rows']}")
            print(f"Total columns: {analysis['total_columns']}")
            print("\nPotentially sensitive columns:")
            for col in analysis['potentially_sensitive_columns']:
                print(f"  - {col}")
            
            # Process the file
            print("\nProcessing file...")
            results = processor.process_file(input_file, output_file)
            
            print("\nProcessing Results:")
            print(f"Total cells: {results['total_cells']}")
            print(f"Cells with PII: {results['cells_with_pii']}")
            print("\nPII types found:")
            for pii_type, count in results['pii_types'].items():
                print(f"  - {pii_type}: {count}")
            print(f"\nOutput file: {results['output_file']}")
            
        except Exception as e:
            print(f"Error: {str(e)}", file=sys.stderr)
            sys.exit(1)
    else:
        print("Usage: python excel_processor.py input.xlsx [output.xlsx]")
        sys.exit(1)